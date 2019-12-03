import openpyxl
import os

all_columns = []
date = ""
load_number = "AAA"
origin_city = "AAA"
origin_state = "AAA"
origin_zip = "AAA"
origin_region = "AAA"
destination_city = "AAA"
destination_state = "AAA"
destination_zip = "AAA"
destination_region = "AAA"
equipment = "AAA"
volume = "AAA"
dat_miles = "AAA"
cust_miles = "AAA"
dat_15 = "AAA"
dat_60 = "AAA"
its_30 = "AAA"
its_60 = "AAA"
rate_avg = "AAA"
year_avg = "AAA"
std_dev = "AAA"

#appends letters A-Z to all columns list
for letter in range(65, 91):
    all_columns.append(chr(letter))

directory = r"C:\Users\djenkins\Desktop\RA_TEST"
all_lanes = r"C:\Users\djenkins\Desktop\all_lanes.xlsx"
master_book = openpyxl.load_workbook(all_lanes)
master_sheet = master_book.active
#master_max_row = master_sheet.max_row

for filename in os.listdir(directory):
    #master_book.save(all_lanes)
    file_ext = filename[-5:]
    print(filename)
    if file_ext == ".xlsx":
        open_file = openpyxl.load_workbook(os.path.join(directory, filename), data_only=True)
        main_sheet = open_file["consolidated data"]
        max_row = main_sheet.max_row + 1
        master_max_row = master_sheet.max_row
        #mapping column for each data point
        for column in all_columns:
            if main_sheet[column + '2'].value == "Load #":
                load_number = column
            if main_sheet[column + '2'].value == "Origin City":
                origin_city = column
            if main_sheet[column + '2'].value == "Origin State":
                origin_state = column
            if main_sheet[column + '2'].value == "Origin Zip Code":
                origin_zip = column
            if main_sheet[column + '2'].value == "Origin Region":
                origin_region = column
            if main_sheet[column + '2'].value == "Destination City":
                destination_city = column
            if main_sheet[column + '2'].value == "Destination State":
                destination_state = column
            if main_sheet[column + '2'].value == "Destination Zip":
                destination_zip = column
            if main_sheet[column + '2'].value == "Destination Region":
                destination_region = column
            if main_sheet[column + '2'].value == "Equipment":
                equipment = column
            if main_sheet[column + '2'].value == "Volume/Year":
                volume = column
            if main_sheet[column + '2'].value == "DAT Miles":
                dat_miles = column
            if main_sheet[column + '2'].value == "CUST MILES":
                cust_miles = column
            if main_sheet[column + '2'].value == "DAT 15 DAY":
                dat_15 = column
            if main_sheet[column + '2'].value == "DAT 60 DAY":
                dat_60 = column
            if main_sheet[column + '2'].value == "ITS 30 DAY":
                its_30 = column
            if main_sheet[column + '2'].value == "ITS 60 DAY":
                its_60 = column
            if main_sheet[column + '2'].value == "60 DAY AVG":
                rate_avg = column
            if main_sheet[column + '2'].value == "ITS YEAR AVG":
                year_avg = column
            if main_sheet[column + '2'].value == "Standard Deviation":
                std_dev = column

            for x in range(master_max_row + 1, max_row + master_max_row - 2):
                master_max_ind = master_max_row - 2
                customer = main_sheet['A1'].value
                customer = customer.split(".")
                master_sheet['a' + str(x)].value = customer[0]
                master_sheet['B' + str(x)].value = main_sheet[load_number + str(x - master_max_ind)].value
                master_sheet['C' + str(x)].value = main_sheet[origin_city + str(x - master_max_ind)].value
                master_sheet['D' + str(x)].value = main_sheet[origin_state + str(x - master_max_ind)].value
                master_sheet['E' + str(x)].value = main_sheet[origin_zip + str(x - master_max_ind)].value
                master_sheet['F' + str(x)].value = main_sheet[origin_region + str(x - master_max_ind)].value
                master_sheet['G' + str(x)].value = main_sheet[destination_city + str(x - master_max_ind)].value
                master_sheet['H' + str(x)].value = main_sheet[destination_state + str(x - master_max_ind)].value
                master_sheet['I' + str(x)].value = main_sheet[destination_zip + str(x - master_max_ind)].value
                master_sheet['J' + str(x)].value = main_sheet[destination_region + str(x - master_max_ind)].value
                master_sheet['K' + str(x)].value = main_sheet[equipment + str(x - master_max_ind)].value
                master_sheet['L' + str(x)].value = main_sheet[volume + str(x - master_max_ind)].value
                master_sheet['M' + str(x)].value = main_sheet[dat_miles + str(x - master_max_ind)].value
                master_sheet['N' + str(x)].value = main_sheet[cust_miles + str(x - master_max_ind)].value
                master_sheet['O' + str(x)].value = main_sheet[dat_15 + str(x - master_max_ind)].value
                master_sheet['P' + str(x)].value = main_sheet[dat_60 + str(x - master_max_ind)].value
                master_sheet['Q' + str(x)].value = main_sheet[its_30 + str(x - master_max_ind)].value
                master_sheet['R' + str(x)].value = main_sheet[its_60 + str(x - master_max_ind)].value
                master_sheet['S' + str(x)].value = main_sheet[rate_avg + str(x - master_max_ind)].value
                master_sheet['T' + str(x)].value = main_sheet[year_avg + str(x - master_max_ind)].value
                master_sheet['U' + str(x)].value = main_sheet[std_dev + str(x - master_max_ind)].value
                master_sheet['V' + str(x)].value = main_sheet['C1'].value
                master_sheet['W' + str(x)].value = main_sheet['A1'].value

master_book.save(all_lanes)
