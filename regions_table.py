import openpyxl
import os

region_table = openpyxl.load_workbook("Region Table.xlsx", data_only=True)
region_sheet = region_table.active
region_columns = ["A", "B"]
region_rows = region_sheet.max_row
column_list = []
origin_city = ""
destination_city = ""
origin_city_index = 0
destination_city_index = 0

#appends letters A-Z to
for letter in range(65, 91):
    column_list.append(chr(letter))

directory = r"C:\Users\djenkins\Desktop\RA_TEST"

for filename in os.listdir(directory):
    file_ext = filename[-5:]
    print(filename)
    if file_ext == ".xlsx":
        open_file = openpyxl.load_workbook(os.path.join(directory, filename), data_only=True)
        #saves spreadsheets as value only deleting all existing formulas
        open_file.save(os.path.join(directory, filename))
        open_file.create_sheet("Regions Table")
        sheet = open_file["Regions Table"]
        main_sheet = open_file["consolidated data"]
        max_row = main_sheet.max_row + 1
        #copies all cells from the region table spreadsheet to the rate analysis region table sheet if origin region has not already been added to the sheet
        if main_sheet['F2']. value != "Origin Region":
            for x in region_columns:
                for y in range(1, region_rows):
                    sheet[x + str(y)].value = region_sheet[x + str(y)].value

            #verifies cell for ORzip and DSzip and creates a new column after DSzip and OR Zip
            for x in column_list:
                if main_sheet[x + '2'].value == "Origin Zip Code":
                    origin_city = x
                    origin_city_index = column_list.index(origin_city)
                    main_sheet.insert_cols(int(origin_city_index + 2))
                    main_sheet[column_list[origin_city_index + 1] + '2'].value = "Origin Region"
                elif main_sheet[x + '2'].value == "Destination Zip":
                    destination_city = x
                    destination_city_index = column_list.index(destination_city)
                    main_sheet.insert_cols(int(destination_city_index + 2))
                    main_sheet[column_list[destination_city_index + 1] + '2'].value = "Destination Region"

            for x in range(3, max_row):
                main_sheet[column_list[column_list.index(origin_city) + 1] + str(x)].value = "=VLOOKUP(LEFT(" + origin_city + str(x) + ", 3), 'Regions Table'!A:B, 2, FALSE)"
                main_sheet[column_list[column_list.index(destination_city) + 1] + str(x)].value = "=VLOOKUP(LEFT(" + destination_city + str(x) + ", 3), 'Regions Table'!A:B, 2, FALSE)"

                #checks for origin zip exceptions
                if len(str(main_sheet[origin_city + str(x)].value)) == 4:
                    main_sheet[origin_city + str(x)].value = "0" + str(main_sheet[origin_city + str(x)].value)
                elif str(main_sheet[origin_city + str(x)].value) == "5":
                    main_sheet[origin_city + str(x)].value = "005"
                elif len(str(main_sheet[origin_city + str(x)].value)) == 3:
                    main_sheet[origin_city + str(x)].value = str(main_sheet[origin_city + str(x)].value) + "00"
                elif len(str(main_sheet[origin_city + str(x)]. value)) == 2:
                    main_sheet[origin_city + str(x)].value = "00" + str(main_sheet[origin_city + str(x)].value) + "00"
                # checks for destination zip exceptions
                if len(str(main_sheet[destination_city + str(x)].value)) == 4:
                    main_sheet[destination_city + str(x)].value = "0" + str(main_sheet[destination_city + str(x)].value)
                elif str(main_sheet[destination_city + str(x)].value) == "5":
                    main_sheet[destination_city + str(x)].value = "005"
                elif len(str(main_sheet[destination_city + str(x)].value)) == 3:
                    main_sheet[destination_city + str(x)].value = str(main_sheet[destination_city + str(x)].value) + "00"
                elif len(str(main_sheet[destination_city + str(x)]. value)) == 2:
                    main_sheet[destination_city + str(x)].value = "00" + str(main_sheet[destination_city + str(x)].value) + "00"

                #checks for canadian zip exceptions
                if len(str(main_sheet[origin_city + str(x)].value)) >= 6:
                    main_sheet[column_list[column_list.index(origin_city) + 1] + str(x)].value = "CANADA"
                if len(str(main_sheet[destination_city + str(x)].value)) >= 6:
                    main_sheet[column_list[column_list.index(destination_city) + 1] + str(x)].value = "CANADA"

            open_file.save(os.path.join(directory, filename))

        else:
            print("Skipped")